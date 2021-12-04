import datetime
import http
import os
from zipfile import ZipFile

import openpyxl
from botocore.exceptions import ClientError

import database
import boto3
from botocore.config import Config


class Exporter:
    def __init__(self, test):
        boto_config = Config(
            region_name='us-east-1',
            signature_version='s3v4',
            retries={
                'max_attempts': 10,
                'mode': 'standard'
            }
        )
        self.test = test
        if not test:
            self.client = boto3.client('s3', config=boto_config)
            self.bucket = os.getenv("EXPORT_BUCKET")

    def export_log(self, database, log_id):
        log = database.get_log(log_id)
        template = database.get_template_internal(log["template"])
        template_headers = template["headers"].keys()
        slides = database.get_slides_internal(log_id)
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = log["name"]
        for i, key in enumerate(template_headers):
            worksheet.cell(1, i+1, key)
        for i, slide in enumerate(slides):
            for j, key in enumerate(template_headers):
                worksheet.cell(2+i, j+1, slide["fields"][key])
        workbook.save(f"{log['name']}.xlsx")

    def export_all_logs(self, database):
        logs = database.get_logs()
        files = []
        for log in logs:
            template = database.get_template_internal(log["template"])
            template_headers = template["headers"].keys()
            slides = database.get_slides_internal(log["id"])
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = log["name"]
            for i, key in enumerate(template_headers):
                worksheet.cell(1, i + 1, key)
            for i, slide in enumerate(slides):
                for j, key in enumerate(template_headers):
                    worksheet.cell(2 + i, j + 1, slide["fields"][key])
            file_name = f"{log['name'].replace(' ', '_')}.xlsx"
            workbook.save(file_name)
            files.append(file_name)
        backup_file_name = f"all_logs.zip"
        with ZipFile(backup_file_name, 'w') as backup:
            for file in files:
                backup.write(file)
        if not self.test:
            try:
                response = self.client.upload_file(backup_file_name, self.bucket, backup_file_name, ExtraArgs={'StorageClass': 'INSTANT_RETRIEVAL'})
                for file in files:
                    os.remove(file)
            except ClientError as e:
                return False
                print(e)
        return f"https://{self.bucket}.s3.amazonaws.com/{backup_file_name}"


    def export_log_date(self, database, log_id, date):
        log = database.get_log(log_id)
        template = database.get_template_internal(log["template"])
        template_headers = template["headers"].keys()
        slides = database.get_slides_internal(log_id)
        filtered_slides = [slide for slide in slides if datetime.date.fromisoformat(slide["created"][:10]) == date]
        filtered_slides = [slide for slide in filtered_slides if slide["submitted"]]
        clean_name = log['name'].replace(' ', '_').replace('\\', '-')
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = clean_name
        for i, key in enumerate(template_headers):
            worksheet.cell(1, i+1, key)
        for i, slide in enumerate(filtered_slides):
            for j, key in enumerate(template_headers):
                worksheet.cell(2+i, j+1, slide["fields"][key])
        file_name = f"{clean_name}-{date.isoformat()}.xlsx"
        workbook.save(file_name)
        if not self.test:
            try:
                response = self.client.upload_file(file_name, self.bucket, file_name)
                os.remove(file_name)
            except ClientError as e:
                print(e)
                return None
            return f"https://{self.bucket}.s3.amazonaws.com/{file_name}"
        else:
            return "url"
